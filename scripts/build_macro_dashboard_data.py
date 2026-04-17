#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
构建“外部宏观数据”驱动的仪表盘数据：

1. 读取指标清单说明表，建立指标目录与优先级映射
2. 读取可用的日度 Excel 数据，映射到系统指标
3. 计算派生指标（期限利差、信用利差、中美利差）
4. 输出前端直接消费的 `data/macro_dashboard.json`
"""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - 环境依赖提醒
    raise SystemExit(
        "缺少 openpyxl。请优先使用项目内 `.venv/bin/python scripts/build_macro_dashboard_data.py` 运行。"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
EXTERNAL_DIR = ROOT / "外部宏观数据"
DATA_DIR = ROOT / "data"
OUTPUT_JSON = DATA_DIR / "macro_dashboard.json"

CATALOG_FILE = EXTERNAL_DIR / "macro_indicator_list.xlsx"
DAILY_GLOB = "宏观经济-日_*.xlsx"
MONTHLY_GLOB = "宏观经济指标_*.xlsx"
LME_COPPER_GLOB = "现货结算价LME铜_*.xlsx"
MANUAL_OVERRIDE_FILE = EXTERNAL_DIR / "manual_macro_series_overrides.json"

SERIES_WINDOW = 40
PERCENTILE_WINDOW = 1250

DERIVED_INDICATOR_IDS = {
    "TERM_SPREAD",
    "CREDIT_SPREAD_AAA_3Y",
    "CREDIT_SPREAD_AA_3Y",
    "CN_US_SPREAD_10Y",
    "M1_M2_GAP",
}

WIDE_TIMESERIES_HEADER_ROW = 2
WIDE_TIMESERIES_DATA_START_ROW = 5


def pick_latest_workbook(pattern: str) -> Path:
    candidates = sorted(EXTERNAL_DIR.glob(pattern))
    if not candidates:
        return EXTERNAL_DIR / pattern.replace("*", "missing")
    return candidates[-1]


DAILY_FILE = pick_latest_workbook(DAILY_GLOB)
MONTHLY_FILE = pick_latest_workbook(MONTHLY_GLOB)
LME_COPPER_FILE = pick_latest_workbook(LME_COPPER_GLOB)

SERIES_NAME_MAP = {
    "DR007": "DR007",
    "R007": "R007",
    "Shibor:3月": "SHIBOR_3M",
    "Shibor:1周": "SHIBOR_1W",
    "中债国债到期收益率:1年": "CGB_1Y",
    "中债国债到期收益率:10年": "CGB_10Y",
    "逆回购:7日:回购利率": "OMO_7D",
    "存款准备金率:大型金融机构": "RRR_BIG",
    "美国:国债收益率:10年": "UST_10Y",
    "美国:国债收益率:2年": "UST_2Y",
    "美国:联邦基金目标利率": "FED_RATE",
    "美国:美元指数": "DXY",
    "中间价:美元兑人民币": "USDCNY_CNB",
    "南华期货:商品指数": "NH_COMMODITY",
    "现货价:Brent原油": "BRENT_OIL",
    "期货结算价(连续):COMEX黄金": "COMEX_GOLD",
    "RJ/CRB商品价格指数": "CRB_INDEX",
    "沪深300指数": "CSI300",
    "中证500指数": "CSI500",
    "中证1000指数": "CSI1000",
    "中债国债总指数(总值)财富指数": "CBA_GOV_WEALTH",
    "中债总指数(总值)财富指数": "CBA_TOTAL_WEALTH",
    "滚动市盈率(TTM):沪深300": "CSI300_PE",
    "恒生指数": "HSI",
    "标准普尔500指数": "SPX",
}

SUPPORT_SERIES = {
    "中债国开债到期收益率:3年": "CDB_3Y",
    "中债企业债到期收益率(AAA):3年": "CB_AAA_3Y",
    "中债企业债到期收益率(AA):3年": "CB_AA_3Y",
}

EXTRA_SERIES_ALIASES = {
    "美国10年国债收益率": "UST_10Y",
    "美国2年国债收益率": "UST_2Y",
    "美元兑人民币中间价": "USDCNY_CNB",
    "沪深300滚动市盈率": "CSI300_PE",
    "沪深300滚动pe": "CSI300_PE",
    "沪深300pe": "CSI300_PE",
    "规模以上工业增加值当月同比": "IVA_YoY",
    "规模以上工业增加值累计同比": "IVA_YTD_YoY",
    "固定资产投资不含农户完成额累计同比": "FAI_YTD_YoY",
    "固定资产投资不含农户完成额制造业累计同比": "MFG_INV_YTD",
    "固定资产投资不含农户完成额基础设施建设投资累计同比": "INFRA_INV_YTD",
    "出口总值美元计价当月同比": "EXPORT_USD_YoY",
    "进口总值美元计价当月同比": "IMPORT_USD_YoY",
    "工业产能利用率当季值": "CAPACITY_UTIL",
    "规模以上工业企业利润总额累计同比": "IND_PROFIT_YTD",
    "cpi非食品当月同比": "CPI_NONFOOD",
    "cpi不包括食品和能源当月同比": "CORE_CPI_YoY",
    "cpi食品当月同比": "CPI_FOOD",
    "ppi生产资料当月同比": "PPI_PRODUCTION",
    "ppi生活资料当月同比": "PPI_LIVING",
    "ppi当月同比": "PPI_YoY",
    "gdp平减指数当季同比": "GDP_DEFLATOR",
    "中期借贷便利mlf操作利率1年期": "MLF_1Y",
    "lpr1年": "LPR_1Y",
    "lpr5年": "LPR_5Y",
    "存款准备金率大型存款类金融机构": "RRR_BIG",
    "m0流通中货币同比": "M0_YoY",
    "m1货币同比": "M1_YoY",
    "m2货币和准货币同比": "M2_YoY",
    "社会融资规模增量当月值": "SOCIAL_FIN_NEW",
    "社会融资规模存量期末同比": "SOCIAL_FIN_STOCK_YoY",
    "社会融资规模增量人民币贷款当月值": "SF_RMB_LOAN",
    "社会融资规模增量企业债券融资当月值": "SF_CORP_BOND",
    "社会融资规模增量政府债券净融资当月值初值": "SF_GOV_BOND",
    "金融机构人民币贷款余额同比": "LOAN_BAL_YoY",
    "金融机构人民币贷款当月增加": "NEW_RMB_LOAN",
    "金融机构人民币贷款当月增加住户中长期": "HH_MLT_LOAN_NEW",
    "金融机构人民币贷款当月增加企事业单位中长期贷款": "CORP_MLT_LOAN_NEW",
    "规模以上工业企业产成品存货期末同比": "FG_INV_YTD_YoY",
    "规模以上工业企业营业收入累计同比": "IND_REVENUE_YTD",
    "停规模以上工业企业主营业务收入累计同比": "IND_MAIN_REV_YTD",
    "规模以上工业企业存货期末同比": "INV_TOTAL_YTD",
    "ppi采掘当月同比": "PPI_MINING",
    "ppi生产资料采掘业": "PPI_MINING",
    "美国供应管理协会ism制造业pmi": "US_ISM_MFG",
    "美国cpi当月同比": "US_CPI_YoY",
    "美国新增非农就业人数当月值季调": "US_NFP",
    "美国失业率16岁及以上季调当月值": "US_UNEMP",
    "现货结算价lme铜": "LME_COPPER",
    "现货结算价:lme铜": "LME_COPPER",
    "lme铜现货结算价": "LME_COPPER",
}

SOURCE_LIBRARY = {
    "nbs_pmi": {
        "title": "国家统计局 PMI 释义",
        "url": "https://www.stats.gov.cn/xxgk/sjfb/zxfb2020/202603/t20260331_1962889.html",
        "note": "国家统计局口径将 PMI 50 视为扩张/收缩临界点。",
    },
    "pbc_social_financing": {
        "title": "人民银行社融存量定义",
        "url": "https://www.pbc.gov.cn/goutongjiaoliu/113456/113469/2025092212554883949/index.html",
        "note": "社融存量定义为一定时期末实体经济从金融体系获得的资金余额，并附带口径调整说明。",
    },
    "pbc_repo_ops": {
        "title": "人民银行公开市场 7 天逆回购",
        "url": "https://www.pbc.gov.cn/zhengcehuobisi/125207/125213/125431/125475/5912133/index.html",
        "note": "公开市场 7 天逆回购操作利率是短端政策利率锚。",
    },
    "chinamoney_repo_fixing": {
        "title": "中国货币网回购定盘利率说明",
        "url": "https://www.chinamoney.com.cn/chinese/bkfrr/",
        "note": "给出 FR007/FDR007 编制方法，说明存款类机构和银银间回购基准口径。",
    },
    "fred_dgs10": {
        "title": "FRED DGS10",
        "url": "https://fred.stlouisfed.org/series/DGS10",
        "note": "美国 10 年国债收益率，日频，来源为美联储 H.15。",
    },
    "fred_dgs2": {
        "title": "FRED DGS2",
        "url": "https://fred.stlouisfed.org/series/DGS2",
        "note": "美国 2 年国债收益率，日频，来源为美联储 H.15。",
    },
    "fred_dxy": {
        "title": "FRED DTWEXBGS",
        "url": "https://fred.stlouisfed.org/series/DTWEXBGS",
        "note": "广义美元指数，日频，来源为美联储 H.10。",
    },
}

THEME_DEFINITIONS = [
    {
        "id": "growth-inflation",
        "title": "增长与通胀",
        "summary_hint": "看 PMI、GDP、CPI/PPI 和美国 ISM，区分经济修复、滞胀压力还是低通胀弱增长。",
        "indicators": ["GDP_YoY", "PMI_MFG", "PMI_NEW_ORDER", "CPI_YoY", "PPI_YoY", "US_ISM_MFG"],
        "source_ids": ["nbs_pmi"],
    },
    {
        "id": "liquidity-rates",
        "title": "流动性与利率",
        "summary_hint": "看政策利率、资金利率和长端利率的相对位置，判断当前是“宽货币确认”还是“利率底部再定价”。",
        "indicators": ["DR007", "OMO_7D", "CGB_10Y", "TERM_SPREAD", "RRR_BIG"],
        "source_ids": ["pbc_repo_ops", "chinamoney_repo_fixing"],
    },
    {
        "id": "credit-external",
        "title": "信用与外部压力",
        "summary_hint": "看信用利差、中美利差、美元和人民币，判断风险偏好修复能否顺利扩散。",
        "indicators": [
            "CREDIT_SPREAD_AAA_3Y",
            "CREDIT_SPREAD_AA_3Y",
            "UST_10Y",
            "CN_US_SPREAD_10Y",
            "DXY",
            "USDCNY_CNB",
        ],
        "source_ids": ["pbc_social_financing", "fred_dgs10", "fred_dxy"],
    },
    {
        "id": "credit-inventory",
        "title": "信用与库存",
        "summary_hint": "看货币活化、社融、企业中长贷和库存/收入，判断信用是否从政策走向实体。",
        "indicators": [
            "M1_YoY",
            "M2_YoY",
            "M1_M2_GAP",
            "SOCIAL_FIN_STOCK_YoY",
            "LOAN_BAL_YoY",
            "CORP_MLT_LOAN_NEW",
            "FG_INV_YTD_YoY",
            "IND_MAIN_REV_YTD",
            "IND_REVENUE_YTD",
        ],
        "source_ids": ["pbc_social_financing"],
    },
    {
        "id": "commodities-inflation",
        "title": "商品与再通胀",
        "summary_hint": "看商品综合、油价和黄金，分辨市场交易的是需求修复、供给冲击，还是避险/再通胀对冲。",
        "indicators": ["NH_COMMODITY", "BRENT_OIL", "LME_COPPER", "COMEX_GOLD", "CRB_INDEX"],
        "source_ids": ["fred_dxy"],
    },
    {
        "id": "risk-assets",
        "title": "风险资产与估值",
        "summary_hint": "看权益、债券财富指数与估值同步变化，识别是“债券单边强”还是“股债共振修复”。",
        "indicators": ["CSI300", "CSI500", "HSI", "SPX", "CBA_TOTAL_WEALTH", "CSI300_PE"],
        "source_ids": ["fred_dgs10"],
    },
]

SCENARIO_TITLES = {
    "growth-up-inflation-down": "增长上行 / 通胀回落",
    "growth-up-inflation-up": "增长上行 / 通胀上行",
    "growth-down-inflation-down": "增长下行 / 通胀下行",
    "growth-down-inflation-up": "增长下行 / 通胀上行",
    "wide-money-tight-credit": "宽货币 / 紧信用",
    "wide-money-wide-credit": "宽货币 / 宽信用",
    "tight-money-wide-credit": "紧货币 / 宽信用",
    "tight-money-tight-credit": "紧货币 / 紧信用",
}

INDICATOR_META = {
    "GDP_YoY": {
        "display_name": "GDP同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_growth",
        "description": "季度总量增长锚，适合和高频景气指标交叉验证。",
    },
    "PMI_MFG": {
        "display_name": "制造业PMI",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "level_50_pmi",
        "description": "景气度核心指标，50 为扩张与收缩的临界点。",
    },
    "PMI_NEW_ORDER": {
        "display_name": "PMI新订单",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "level_50_pmi",
        "description": "更前瞻的需求代理指标，常先于生产和库存变化。",
    },
    "CPI_YoY": {
        "display_name": "CPI同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "inflation_level",
        "description": "居民通胀主指标，反映终端需求和食品等价格扰动。",
    },
    "CORE_CPI_YoY": {
        "display_name": "核心CPI同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "inflation_level",
        "description": "剔除食品和能源后更能反映内生需求。",
    },
    "PPI_YoY": {
        "display_name": "PPI同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "inflation_level",
        "description": "上游价格和工业利润传导的关键指标。",
    },
    "US_ISM_MFG": {
        "display_name": "美国ISM制造业PMI",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "level_50_pmi",
        "description": "海外制造业景气锚，适合观察中美库存与周期共振。",
    },
    "DR007": {
        "display_name": "DR007",
        "delta_mode": "bp",
        "value_format": "percent",
        "signal_mode": "low_is_loose",
        "description": "银行间流动性黄金指标，观察短端资金面是否贴着政策利率运行。",
    },
    "OMO_7D": {
        "display_name": "7天逆回购利率",
        "delta_mode": "bp",
        "value_format": "percent",
        "signal_mode": "policy_rate",
        "description": "人民银行短端政策利率锚，用来判断货币政策操作区间。",
    },
    "CGB_10Y": {
        "display_name": "中国10Y国债",
        "delta_mode": "bp",
        "value_format": "percent",
        "signal_mode": "low_is_loose",
        "description": "长端利率常同时反映名义增长、通胀和政策预期。",
    },
    "TERM_SPREAD": {
        "display_name": "期限利差(10Y-1Y)",
        "delta_mode": "bp",
        "value_format": "bp",
        "signal_mode": "high_is_steep",
        "description": "10 年减 1 年期限利差，用于观察曲线走陡/走平和增长预期变化。",
    },
    "RRR_BIG": {
        "display_name": "大型行存准",
        "delta_mode": "bp",
        "value_format": "percent",
        "signal_mode": "policy_rate",
        "description": "数量型货币工具，更新不规则，但政策信号强。",
    },
    "M1_YoY": {
        "display_name": "M1同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_credit",
        "description": "企业活期存款增速，常作为经济活跃度和风险偏好的高敏感代理。",
    },
    "M2_YoY": {
        "display_name": "M2同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_credit",
        "description": "广义货币增速，用于观察总量流动性环境。",
    },
    "M1_M2_GAP": {
        "display_name": "M1-M2剪刀差",
        "delta_mode": "pct",
        "value_format": "number_2",
        "signal_mode": "gap_positive_is_active",
        "description": "M1 与 M2 的增速差，常用于观察货币活化程度。",
    },
    "SOCIAL_FIN_STOCK_YoY": {
        "display_name": "社融存量同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_credit",
        "description": "信用环境核心总量指标，观察信用扩张是否持续。",
    },
    "SOCIAL_FIN_NEW": {
        "display_name": "社融当月新增",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "higher_is_credit",
        "description": "信用脉冲高波动指标，适合做阶段性拐点观察。",
    },
    "LOAN_BAL_YoY": {
        "display_name": "贷款余额同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_credit",
        "description": "银行信贷核心指标，观察信用派生是否顺畅。",
    },
    "CORP_MLT_LOAN_NEW": {
        "display_name": "企业中长贷新增",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "higher_is_credit",
        "description": "企业资本开支与实体确认的关键领先指标。",
    },
    "CREDIT_SPREAD_AAA_3Y": {
        "display_name": "AAA利差(3Y)",
        "delta_mode": "bp",
        "value_format": "bp",
        "signal_mode": "low_is_tightening_spread",
        "description": "AAA 企业债相对国开债利差，反映高等级信用风险溢价。",
    },
    "CREDIT_SPREAD_AA_3Y": {
        "display_name": "AA利差(3Y)",
        "delta_mode": "bp",
        "value_format": "bp",
        "signal_mode": "low_is_tightening_spread",
        "description": "AA 企业债相对国开债利差，更敏感地反映低等级风险偏好。",
    },
    "UST_10Y": {
        "display_name": "美国10Y国债",
        "delta_mode": "bp",
        "value_format": "percent",
        "signal_mode": "high_is_tight",
        "description": "全球无风险利率锚，影响中美利差、成长股估值和美元方向。",
    },
    "CN_US_SPREAD_10Y": {
        "display_name": "中美10Y利差",
        "delta_mode": "bp",
        "value_format": "bp",
        "signal_mode": "cn_us_spread",
        "description": "中国 10Y 国债减美国 10Y 国债，反映内外利率错位和汇率压力。",
    },
    "DXY": {
        "display_name": "美元指数",
        "delta_mode": "pct",
        "value_format": "number_2",
        "signal_mode": "currency_lower_is_easier",
        "description": "外部流动性与全球风险偏好的重要代理变量。",
    },
    "USDCNY_CNB": {
        "display_name": "美元兑人民币中间价",
        "delta_mode": "pct",
        "value_format": "number_4",
        "signal_mode": "currency_lower_is_easier",
        "description": "人民币汇率压力指标，联动中美利差与美元周期。",
    },
    "NH_COMMODITY": {
        "display_name": "南华商品指数",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "higher_is_reflation",
        "description": "国内商品综合代理，兼顾需求修复和供给扰动线索。",
    },
    "BRENT_OIL": {
        "display_name": "Brent原油",
        "delta_mode": "pct",
        "value_format": "currency_2",
        "signal_mode": "higher_is_reflation",
        "description": "全球通胀驱动和供给冲击信号，对通胀交易很敏感。",
    },
    "LME_COPPER": {
        "display_name": "LME铜",
        "delta_mode": "pct",
        "value_format": "currency_2",
        "signal_mode": "higher_is_reflation",
        "description": "博士铜，高敏感反映全球制造业和再通胀预期。",
    },
    "COMEX_GOLD": {
        "display_name": "COMEX黄金",
        "delta_mode": "pct",
        "value_format": "currency_0",
        "signal_mode": "higher_is_hedge",
        "description": "避险与再通胀对冲资产，常与美元和真实利率联动。",
    },
    "CRB_INDEX": {
        "display_name": "CRB商品指数",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "higher_is_reflation",
        "description": "全球商品价格综合观察窗，用于交叉验证再通胀交易。",
    },
    "FG_INV_YTD_YoY": {
        "display_name": "产成品库存累计同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "inventory_cycle",
        "description": "库存周期主指标，需和收入或订单一起看。",
    },
    "IND_REVENUE_YTD": {
        "display_name": "工业营收累计同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_growth",
        "description": "工业需求代理，适合与库存同比联动观察。",
    },
    "IND_MAIN_REV_YTD": {
        "display_name": "主营收入累计同比",
        "delta_mode": "pct",
        "value_format": "percent",
        "signal_mode": "higher_is_growth",
        "description": "2012 年前工业收入代理，可与库存周期旧口径衔接。",
    },
    "CSI300": {
        "display_name": "沪深300",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "higher_is_risk_on",
        "description": "A 股宽基风险偏好代理，适合和信用/估值信号一起看。",
    },
    "CSI500": {
        "display_name": "中证500",
        "delta_mode": "pct",
        "value_format": "number_1",
        "signal_mode": "higher_is_risk_on",
        "description": "中盘风格代理，常对信用修复和成长风格更敏感。",
    },
    "HSI": {
        "display_name": "恒生指数",
        "delta_mode": "pct",
        "value_format": "number_2",
        "signal_mode": "higher_is_risk_on",
        "description": "港股风险偏好代理，受国内政策和外部美元环境双重影响。",
    },
    "SPX": {
        "display_name": "标普500",
        "delta_mode": "pct",
        "value_format": "number_2",
        "signal_mode": "higher_is_risk_on",
        "description": "海外权益风险偏好代理，需要和美债利率一起看。",
    },
    "CBA_TOTAL_WEALTH": {
        "display_name": "中债总财富",
        "delta_mode": "pct",
        "value_format": "number_2",
        "signal_mode": "higher_is_bond_strong",
        "description": "中国债券整体表现代理，利率与信用环境变化会共同影响走势。",
    },
    "CSI300_PE": {
        "display_name": "沪深300滚动PE",
        "delta_mode": "pct",
        "value_format": "number_2",
        "signal_mode": "valuation_lower_is_cheaper",
        "description": "估值分位参考，用来区分“价格涨了”和“估值已经过热”。",
    },
}

MISSING_PRIORITY_IDS = {
    "GDP_YoY",
    "PMI_MFG",
    "PMI_NEW_ORDER",
    "FAI_YTD_YoY",
    "M1_YoY",
    "M2_YoY",
    "M1_M2_GAP",
    "SOCIAL_FIN_STOCK_YoY",
    "SOCIAL_FIN_NEW",
    "LOAN_BAL_YoY",
    "CORP_MLT_LOAN_NEW",
    "FG_INV_YTD_YoY",
    "IND_MAIN_REV_YTD",
    "CPI_YoY",
    "CORE_CPI_YoY",
    "PPI_YoY",
    "US_ISM_MFG",
}


def iso_date(value: datetime) -> str:
    return value.strftime("%Y-%m-%d")


def format_number(value: float, digits: int = 1) -> str:
    return f"{value:,.{digits}f}"


def value_percentile(series: List[float], value: float) -> Tuple[float, str]:
    if not series:
        return 0.0, "—"
    less_or_equal = sum(item <= value for item in series)
    percentile = less_or_equal / len(series)
    window_label = "近5年" if len(series) >= min(PERCENTILE_WINDOW, 250) else "全样本"
    return percentile, f"{window_label} {round(percentile * 100)}% 分位"


def select_percentile_window(values: List[float]) -> List[float]:
    if len(values) > PERCENTILE_WINDOW:
        return values[-PERCENTILE_WINDOW:]
    return values


def format_value(value: float, fmt: str) -> str:
    if fmt == "percent":
        return f"{value:.2f}%"
    if fmt == "bp":
        return f"{value:.1f}bp"
    if fmt == "currency_2":
        return f"${value:,.2f}"
    if fmt == "currency_0":
        return f"${value:,.0f}"
    if fmt == "number_4":
        return f"{value:.4f}"
    if fmt == "number_2":
        return f"{value:,.2f}"
    if fmt == "number_1":
        return f"{value:,.1f}"
    return format_number(value, 2)


def format_delta(current: float, previous: Optional[float], mode: str) -> str:
    if previous is None:
        return "—"
    delta = current - previous
    if mode == "bp":
        return f"{delta * 100:+.1f}bp"
    if previous == 0:
        return "—"
    change = (current / previous - 1) * 100
    return f"{change:+.1f}%"


def latest_change(series: List[Tuple[datetime, float]], offset: int) -> Optional[float]:
    if len(series) <= offset:
        return None
    return series[-(offset + 1)][1]


def latest_observation(series_map: Dict[str, List[Tuple[datetime, float]]], indicator_id: str) -> Tuple[Optional[datetime], Optional[float]]:
    series = series_map.get(indicator_id) or []
    if not series:
        return None, None
    return series[-1]


def latest_value(series_map: Dict[str, List[Tuple[datetime, float]]], indicator_id: str) -> Optional[float]:
    return latest_observation(series_map, indicator_id)[1]


def previous_value(series_map: Dict[str, List[Tuple[datetime, float]]], indicator_id: str, periods: int) -> Optional[float]:
    series = series_map.get(indicator_id) or []
    if len(series) <= periods:
        return None
    return series[-(periods + 1)][1]


def recent_indicator(series_map: Dict[str, List[Tuple[datetime, float]]], indicator_id: str, max_stale_days: int = 70) -> Optional[float]:
    latest_date, latest = latest_observation(series_map, indicator_id)
    if latest_date is None or latest is None:
        return None
    latest_candidates = [
        series[-1][0]
        for series in series_map.values()
        if series and isinstance(series[-1][0], datetime)
    ]
    if not latest_candidates:
        return latest
    max_date = max(latest_candidates)
    if (max_date.date() - latest_date.date()).days > max_stale_days:
        return None
    return latest


def safe_mean(values: List[Optional[float]]) -> Optional[float]:
    filtered = [value for value in values if value is not None]
    if not filtered:
        return None
    return sum(filtered) / len(filtered)


def compare_momentum(current: Optional[float], previous: Optional[float], threshold: float) -> str:
    if current is None or previous is None:
        return "flat"
    change = current - previous
    if change > threshold:
        return "up"
    if change < -threshold:
        return "down"
    return "flat"


def build_signal_label(mode: str, latest: float, percentile: float, delta20: Optional[float]) -> Tuple[str, str]:
    if mode == "level_50_pmi":
        if latest >= 50:
            return "景气扩张", "supportive"
        if latest >= 48:
            return "收缩边缘", "watch"
        return "景气收缩", "pressured"
    if mode == "higher_is_growth":
        if percentile >= 0.8:
            return "增长偏强", "supportive"
        if percentile <= 0.2:
            return "增长偏弱", "pressured"
        return "中性", "neutral"
    if mode == "inflation_level":
        if percentile >= 0.8:
            return "通胀抬升", "watch"
        if percentile <= 0.2:
            return "通胀偏低", "neutral"
        return "中性", "neutral"
    if mode == "low_is_loose":
        if percentile <= 0.2:
            return "资金偏松", "supportive"
        if percentile >= 0.8:
            return "资金偏紧", "pressured"
        return ("边际走松", "supportive") if (delta20 is not None and delta20 < 0) else ("中性", "neutral")
    if mode == "policy_rate":
        return "政策锚", "watch"
    if mode == "higher_is_credit":
        if percentile >= 0.8:
            return "信用偏强", "supportive"
        if percentile <= 0.2:
            return "信用偏弱", "pressured"
        return "中性", "neutral"
    if mode == "gap_positive_is_active":
        if latest > 0:
            return "货币活化", "supportive"
        if latest < -1:
            return "活化偏弱", "pressured"
        return "中性", "neutral"
    if mode == "high_is_steep":
        if latest < 0:
            return "曲线倒挂", "pressured"
        if percentile >= 0.8:
            return "曲线走陡", "supportive"
        if percentile <= 0.2:
            return "曲线偏平", "watch"
        return "中性", "neutral"
    if mode == "low_is_tightening_spread":
        if percentile <= 0.2:
            return "利差收敛", "supportive"
        if percentile >= 0.8:
            return "利差走扩", "pressured"
        return "中性", "neutral"
    if mode == "high_is_tight":
        if percentile >= 0.8:
            return "外部偏紧", "pressured"
        if percentile <= 0.2:
            return "外部缓和", "supportive"
        return "中性", "neutral"
    if mode == "cn_us_spread":
        if latest < 0:
            return "利差倒挂", "pressured"
        if percentile >= 0.8:
            return "利差修复", "supportive"
        return "中性", "neutral"
    if mode == "currency_lower_is_easier":
        if percentile <= 0.2:
            return "外压缓和", "supportive"
        if percentile >= 0.8:
            return "外压抬升", "pressured"
        return "中性", "neutral"
    if mode == "higher_is_reflation":
        if percentile >= 0.8:
            return "再通胀抬头", "watch"
        if percentile <= 0.2:
            return "需求偏弱", "neutral"
        return "中性", "neutral"
    if mode == "inventory_cycle":
        if delta20 is not None and delta20 > 0 and percentile >= 0.6:
            return "库存抬升", "watch"
        if delta20 is not None and delta20 < 0 and percentile <= 0.4:
            return "去库延续", "neutral"
        return "中性", "neutral"
    if mode == "higher_is_hedge":
        if percentile >= 0.8:
            return "避险升温", "watch"
        if percentile <= 0.2:
            return "避险降温", "neutral"
        return "中性", "neutral"
    if mode == "higher_is_risk_on":
        if percentile >= 0.8:
            return "风险偏好回暖", "supportive"
        if percentile <= 0.2:
            return "风险偏好偏弱", "pressured"
        return "中性", "neutral"
    if mode == "higher_is_bond_strong":
        if percentile >= 0.8:
            return "债券偏强", "supportive"
        if percentile <= 0.2:
            return "债券承压", "pressured"
        return "中性", "neutral"
    if mode == "valuation_lower_is_cheaper":
        if percentile <= 0.2:
            return "估值偏低", "supportive"
        if percentile >= 0.8:
            return "估值偏高", "watch"
        return "中性", "neutral"
    return "中性", "neutral"


def signal_observation(percentile_text: str, description: str, latest_date: str, staleness_days: int) -> str:
    freshness = "高频更新" if staleness_days <= 3 else f"最近一次更新距今 {staleness_days} 天"
    return f"{percentile_text}，{freshness}。{description}最新观测日为 {latest_date}。"


def normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    for token in [" ", "\n", "\t", ":", "：", "-", "_", "/", "（", "）", "(", ")", ",", "，", ".", "%"]:
        text = text.replace(token, "")
    return text


def resolve_latest_workbook(pattern: str) -> Optional[Path]:
    candidates = sorted(EXTERNAL_DIR.glob(pattern))
    if not candidates:
        return None
    return max(candidates, key=lambda path: (path.stat().st_mtime, path.name))


def coerce_datetime(value: object) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m", "%Y%m%d", "%Y%m"):
            try:
                parsed = datetime.strptime(raw, fmt)
                if fmt in ("%Y-%m", "%Y/%m", "%Y%m"):
                    return parsed.replace(day=1)
                return parsed
            except ValueError:
                continue
    return None


def coerce_float(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def build_catalog_alias_map(catalog: Dict[str, dict]) -> Dict[str, str]:
    aliases = {}
    for indicator_id, item in catalog.items():
        for raw in [indicator_id, item["name"], item.get("wind_code", "")]:
            normalized = normalize_text(raw)
            if normalized:
                aliases[normalized] = indicator_id
        aliases.setdefault(normalize_text(item["name"]).replace("同比", "yoy"), indicator_id)
    for raw, indicator_id in EXTRA_SERIES_ALIASES.items():
        aliases[normalize_text(raw)] = indicator_id
    return aliases


def inspect_source_workbook(path: Optional[Path]) -> dict:
    if path is None:
        return {
            "status": "missing",
            "summary": "文件不存在。",
            "details": [],
            "layout": "missing",
            "file_name": "",
        }
    workbook_formula = load_workbook(path, read_only=False, data_only=False)
    sheet_formula = workbook_formula[workbook_formula.sheetnames[0]]
    workbook_value = load_workbook(path, read_only=True, data_only=True)
    sheet_value = workbook_value[workbook_value.sheetnames[0]]

    first_formula = sheet_formula["A1"].value
    if (
        isinstance(first_formula, str)
        and first_formula.startswith("=HX_IFIND_EDB")
        and sheet_formula.max_row == 1
        and sheet_formula.max_column == 1
    ):
        return {
            "status": "formula_only",
            "summary": "只有 iFinD 公式锚点，没有缓存值。",
            "details": [
                "工作簿只有 A1 一个公式单元格。",
                "需要在装有 iFinD 插件的 Excel 中刷新并另存，确保结果值真正写入工作簿。",
            ],
            "layout": "formula_only",
            "file_name": path.name,
        }

    rows = list(sheet_value.iter_rows(values_only=True))
    if len(rows) >= WIDE_TIMESERIES_DATA_START_ROW:
        header_row = rows[WIDE_TIMESERIES_HEADER_ROW - 1]
        data_rows = rows[WIDE_TIMESERIES_DATA_START_ROW - 1 :]
        header_count = sum(1 for item in header_row[1:] if item)
        date_hits = sum(1 for row in data_rows[:5] if row and coerce_datetime(row[0]))
        if (header_count >= 2 or (sheet_value.max_column == 2 and header_count >= 1)) and date_hits >= 1:
            return {
                "status": "ready",
                "summary": "识别为宽表时间序列工作簿。",
                "details": [f"{sheet_value.max_row} 行 × {sheet_value.max_column} 列。"],
                "layout": "wide_timeseries",
                "file_name": path.name,
            }

    if rows:
        headers = [normalize_text(item) for item in rows[0] if item is not None]
        if headers and (
            ("日期" in headers or "date" in headers)
            and ("指标名称" in headers or "indicator" in headers or "指标" in headers)
        ):
            return {
                "status": "ready",
                "summary": "识别为长表工作簿。",
                "details": [f"{sheet_value.max_row} 行 × {sheet_value.max_column} 列。"],
                "layout": "long_table",
                "file_name": path.name,
            }

    return {
        "status": "unknown",
        "summary": "文件存在，但未识别出可解析布局。",
        "details": [f"{sheet_value.max_row} 行 × {sheet_value.max_column} 列。"],
        "layout": "unknown",
        "file_name": path.name,
    }


def inspect_monthly_workbook() -> dict:
    status = inspect_source_workbook(MONTHLY_FILE)
    if status.get("status") == "missing":
        status["summary"] = "月/季频工作簿不存在。"
    return status


def read_catalog() -> Dict[str, dict]:
    workbook = load_workbook(CATALOG_FILE, read_only=True, data_only=True)
    sheet = workbook["指标清单"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    catalog = {}
    for row in rows:
        if not any(row):
            continue
        indicator_id = str(row[3]).strip()
        catalog[indicator_id] = {
            "order": row[0],
            "dimension": str(row[1]).strip(),
            "name": str(row[2]).strip(),
            "id": indicator_id,
            "frequency": str(row[4]).strip(),
            "agency": str(row[5]).strip(),
            "wind_code": "" if row[6] is None else str(row[6]).strip(),
            "start_year": "" if row[7] is None else str(row[7]).strip(),
            "unit": "" if row[8] is None else str(row[8]).strip(),
            "frameworks": str(row[9]).strip(),
            "priority": str(row[10]).strip(),
            "note": "" if row[11] is None else str(row[11]).strip(),
        }
    return catalog


def parse_wide_timeseries_workbook(
    path: Path,
    resolver,
) -> Tuple[Dict[str, List[Tuple[datetime, float]]], Optional[datetime]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < WIDE_TIMESERIES_DATA_START_ROW:
        return {}, None

    header_row = rows[WIDE_TIMESERIES_HEADER_ROW - 1]
    data_rows = rows[WIDE_TIMESERIES_DATA_START_ROW - 1 :]
    mapped: Dict[str, Dict[datetime, float]] = defaultdict(dict)
    latest_sheet_date: Optional[datetime] = None

    for row in data_rows:
        date_value = coerce_datetime(row[0] if row else None)
        if not date_value:
            continue
        latest_sheet_date = max(latest_sheet_date, date_value) if latest_sheet_date else date_value
        for index, source_name in enumerate(header_row[1:], start=1):
            if source_name is None:
                continue
            series_id = resolver(source_name)
            if not series_id or index >= len(row):
                continue
            numeric = coerce_float(row[index])
            if numeric is None:
                continue
            mapped[series_id][date_value] = numeric

    normalized = {
        key: sorted(values.items(), key=lambda item: item[0])
        for key, values in mapped.items()
    }
    return normalized, latest_sheet_date


def parse_long_table_workbook(
    path: Path,
    alias_map: Dict[str, str],
) -> Tuple[Dict[str, List[Tuple[datetime, float]]], Optional[datetime]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}, None

    header_row = [normalize_text(item) for item in rows[0]]
    date_index = next((idx for idx, item in enumerate(header_row) if item in {"日期", "date", "时间", "交易日期"}), None)
    name_index = next((idx for idx, item in enumerate(header_row) if item in {"指标名称", "indicator", "指标", "name", "series"}), None)
    value_index = next((idx for idx, item in enumerate(header_row) if item in {"值", "value", "数值", "最新值"}), None)
    if date_index is None or name_index is None or value_index is None:
        return {}, None

    mapped: Dict[str, Dict[datetime, float]] = defaultdict(dict)
    latest_sheet_date: Optional[datetime] = None
    for row in rows[1:]:
        if row is None:
            continue
        date_value = coerce_datetime(row[date_index] if date_index < len(row) else None)
        name_value = row[name_index] if name_index < len(row) else None
        numeric = coerce_float(row[value_index] if value_index < len(row) else None)
        if not date_value or numeric is None:
            continue
        series_id = alias_map.get(normalize_text(name_value))
        if not series_id:
            continue
        latest_sheet_date = max(latest_sheet_date, date_value) if latest_sheet_date else date_value
        mapped[series_id][date_value] = numeric

    normalized = {
        key: sorted(values.items(), key=lambda item: item[0])
        for key, values in mapped.items()
    }
    return normalized, latest_sheet_date


def read_workbook_series(
    path: Optional[Path],
    inspection: dict,
    alias_map: Dict[str, str],
    explicit_name_map: Optional[Dict[str, str]] = None,
) -> Tuple[Dict[str, List[Tuple[datetime, float]]], Optional[datetime]]:
    if path is None or inspection["status"] != "ready":
        return {}, None

    explicit = explicit_name_map or {}

    def resolver(name: object) -> Optional[str]:
        text = str(name or "").strip()
        if text in explicit:
            return explicit[text]
        normalized = normalize_text(text)
        if normalized in alias_map:
            return alias_map[normalized]
        if normalized in explicit:
            return explicit[normalized]
        return None

    if inspection["layout"] == "wide_timeseries":
        return parse_wide_timeseries_workbook(path, resolver)
    if inspection["layout"] == "long_table":
        return parse_long_table_workbook(path, alias_map)
    return {}, None


def read_manual_override_series(
    path: Path,
    catalog: Dict[str, dict],
) -> Tuple[Dict[str, List[Tuple[datetime, float]]], dict]:
    if not path.exists():
        return {}, {
            "status": "missing",
            "summary": "未提供手工覆盖文件。",
            "details": [],
            "layout": "json",
            "file_name": "",
        }

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        return {}, {
            "status": "invalid",
            "summary": "手工覆盖文件 JSON 格式错误。",
            "details": [str(exc)],
            "layout": "json",
            "file_name": path.name,
        }

    raw_series = payload.get("series") or {}
    parsed: Dict[str, Dict[datetime, float]] = defaultdict(dict)
    unknown_ids = []
    invalid_rows = 0
    for indicator_id, records in raw_series.items():
        if indicator_id not in catalog:
            unknown_ids.append(indicator_id)
            continue
        if not isinstance(records, list):
            invalid_rows += 1
            continue
        for item in records:
            if not isinstance(item, dict):
                invalid_rows += 1
                continue
            date_value = coerce_datetime(item.get("date"))
            numeric = coerce_float(item.get("value"))
            if not date_value or numeric is None:
                invalid_rows += 1
                continue
            parsed[indicator_id][date_value] = numeric

    normalized = {
        key: sorted(values.items(), key=lambda item: item[0])
        for key, values in parsed.items()
        if values
    }

    details = []
    if unknown_ids:
        details.append(f"忽略未知指标 ID: {', '.join(sorted(unknown_ids))}")
    if invalid_rows:
        details.append(f"忽略无效记录 {invalid_rows} 条。")
    if normalized:
        details.insert(0, f"已读取 {len(normalized)} 条手工覆盖序列。")
        status = "ready"
        summary = "手工覆盖文件已生效。"
    else:
        status = "empty"
        summary = "手工覆盖文件存在，但当前没有有效数据。"

    return normalized, {
        "status": status,
        "summary": summary,
        "details": details,
        "layout": "json",
        "file_name": path.name,
    }


def intersect_series(
    left: List[Tuple[datetime, float]],
    right: List[Tuple[datetime, float]],
    transform,
) -> List[Tuple[datetime, float]]:
    right_map = {date: value for date, value in right}
    series = []
    for date, value in left:
        if date not in right_map:
            continue
        series.append((date, transform(value, right_map[date])))
    return series


def build_derived_series(series_map: Dict[str, List[Tuple[datetime, float]]]) -> Dict[str, List[Tuple[datetime, float]]]:
    derived = {}
    cgb_10y = series_map.get("CGB_10Y", [])
    cgb_1y = series_map.get("CGB_1Y", [])
    ust_10y = series_map.get("UST_10Y", [])
    cdb_3y = series_map.get("CDB_3Y", [])
    aaa_3y = series_map.get("CB_AAA_3Y", [])
    aa_3y = series_map.get("CB_AA_3Y", [])
    m1_yoy = series_map.get("M1_YoY", [])
    m2_yoy = series_map.get("M2_YoY", [])

    if cgb_10y and cgb_1y:
        derived["TERM_SPREAD"] = intersect_series(cgb_10y, cgb_1y, lambda a, b: (a - b) * 100)
    if aaa_3y and cdb_3y:
        derived["CREDIT_SPREAD_AAA_3Y"] = intersect_series(aaa_3y, cdb_3y, lambda a, b: (a - b) * 100)
    if aa_3y and cdb_3y:
        derived["CREDIT_SPREAD_AA_3Y"] = intersect_series(aa_3y, cdb_3y, lambda a, b: (a - b) * 100)
    if cgb_10y and ust_10y:
        derived["CN_US_SPREAD_10Y"] = intersect_series(cgb_10y, ust_10y, lambda a, b: (a - b) * 100)
    if m1_yoy and m2_yoy:
        derived["M1_M2_GAP"] = intersect_series(m1_yoy, m2_yoy, lambda a, b: a - b)
    return derived


def build_signal_payload(
    indicator_id: str,
    catalog: Dict[str, dict],
    series_map: Dict[str, List[Tuple[datetime, float]]],
    latest_sheet_date: datetime,
) -> Optional[dict]:
    series = series_map.get(indicator_id)
    meta = INDICATOR_META.get(indicator_id)
    entry = catalog.get(indicator_id)
    if not series or not meta or not entry:
        return None

    latest_date, latest_value = series[-1]
    values = [value for _, value in series]
    percentile, percentile_text = value_percentile(select_percentile_window(values), latest_value)
    previous_5 = latest_change(series, 5)
    previous_20 = latest_change(series, 20)
    previous_250 = latest_change(series, 250)
    signal_label, signal_tone = build_signal_label(
        meta["signal_mode"],
        latest_value,
        percentile,
        None if previous_20 is None else latest_value - previous_20,
    )
    staleness_days = (latest_sheet_date.date() - latest_date.date()).days
    return {
        "id": indicator_id,
        "display_name": meta["display_name"],
        "catalog_name": entry["name"],
        "dimension": entry["dimension"],
        "priority": entry["priority"],
        "frameworks": entry["frameworks"],
        "latest_value": format_value(latest_value, meta["value_format"]),
        "latest_numeric": round(latest_value, 6),
        "latest_date": iso_date(latest_date),
        "staleness_days": staleness_days,
        "change_5d": format_delta(latest_value, previous_5, meta["delta_mode"]),
        "change_20d": format_delta(latest_value, previous_20, meta["delta_mode"]),
        "change_250d": format_delta(latest_value, previous_250, meta["delta_mode"]),
        "percentile_text": percentile_text,
        "signal_label": signal_label,
        "signal_tone": signal_tone,
        "description": meta["description"],
        "observation": signal_observation(
            percentile_text,
            meta["description"],
            iso_date(latest_date),
            staleness_days,
        ),
        "sparkline": [round(value, 4) for value in values[-SERIES_WINDOW:]],
    }


def theme_summary(theme_id: str, signal_map: Dict[str, dict]) -> str:
    if theme_id == "growth-inflation":
        pmi = signal_map.get("PMI_MFG")
        order = signal_map.get("PMI_NEW_ORDER")
        cpi = signal_map.get("CPI_YoY")
        ppi = signal_map.get("PPI_YoY")
        if pmi and cpi:
            order_text = f"，新订单 {order['latest_value']}" if order else ""
            ppi_text = f"，PPI {ppi['latest_value']}" if ppi else ""
            return (
                f"制造业 PMI 为 {pmi['latest_value']}{order_text}，"
                f"CPI 为 {cpi['latest_value']}{ppi_text}；可直接区分修复、低通胀或再通胀。"
            )
    if theme_id == "liquidity-rates":
        dr007 = signal_map.get("DR007")
        cgb = signal_map.get("CGB_10Y")
        curve = signal_map.get("TERM_SPREAD")
        if dr007 and cgb and curve:
            return (
                f"DR007 为 {dr007['latest_value']}，{dr007['signal_label']}；"
                f"10Y 国债在 {cgb['latest_value']} 的低位附近运行，"
                f"{curve['signal_label']}。"
            )
    if theme_id == "credit-external":
        aaa = signal_map.get("CREDIT_SPREAD_AAA_3Y")
        spread = signal_map.get("CN_US_SPREAD_10Y")
        dxy = signal_map.get("DXY")
        if aaa and spread and dxy:
            return (
                f"AAA 利差 {aaa['signal_label']}，中美 10Y 利差为 {spread['latest_value']}；"
                f"美元指数当前 {dxy['latest_value']}，{dxy['signal_label']}。"
            )
    if theme_id == "credit-inventory":
        sf = signal_map.get("SOCIAL_FIN_STOCK_YoY")
        corp = signal_map.get("CORP_MLT_LOAN_NEW")
        gap = signal_map.get("M1_M2_GAP")
        inv = signal_map.get("FG_INV_YTD_YoY")
        if sf and gap:
            corp_text = f"，企业中长贷 {corp['latest_value']}" if corp else ""
            inv_text = f"，库存 {inv['latest_value']}" if inv else ""
            return (
                f"社融存量同比 {sf['latest_value']}，M1-M2 剪刀差 {gap['latest_value']}{corp_text}{inv_text}；"
                "核心在于信用是否已传导到实体与库存。"
            )
    if theme_id == "commodities-inflation":
        nh = signal_map.get("NH_COMMODITY")
        brent = signal_map.get("BRENT_OIL")
        copper = signal_map.get("LME_COPPER")
        gold = signal_map.get("COMEX_GOLD")
        if nh and brent and copper and gold:
            return (
                f"南华商品指数位于 {nh['percentile_text']}，油价 {brent['latest_value']}，"
                f"LME铜 {copper['latest_value']}，黄金 {gold['latest_value']}；"
                "市场同时在交易再通胀和避险。"
            )
    if theme_id == "risk-assets":
        csi = signal_map.get("CSI300")
        bond = signal_map.get("CBA_TOTAL_WEALTH")
        pe = signal_map.get("CSI300_PE")
        if csi and bond and pe:
            return (
                f"沪深300 为 {csi['latest_value']}，中债总财富指数为 {bond['latest_value']}；"
                f"股债阶段性同步偏强，但估值仍处 {pe['percentile_text']}。"
            )
    return "当前主题暂无足够数据生成摘要。"


def build_coverage(catalog: Dict[str, dict], available_catalog_ids: Iterable[str], monthly_status: dict) -> dict:
    available_set = set(available_catalog_ids)
    grouped = defaultdict(list)
    for item in catalog.values():
        grouped[item["dimension"]].append(item)

    dimensions = []
    for dimension, items in grouped.items():
        total = len(items)
        p0_total = sum(1 for item in items if item["priority"] == "P0")
        mapped = sum(1 for item in items if item["id"] in available_set)
        mapped_p0 = sum(1 for item in items if item["id"] in available_set and item["priority"] == "P0")
        coverage_ratio = 0 if total == 0 else mapped / total
        if mapped:
            comment = (
                "当前维度已有日频与月/季频混合覆盖。"
                if dimension in {"增长", "通胀", "信用", "库存"} and monthly_status["status"] == "ready"
                else "当前维度已有可直接读取的数据。"
            )
        else:
            comment = "当前维度暂无可直接读取的导出值，主要受月/季频工作簿未落盘影响。"
        dimensions.append(
            {
                "dimension": dimension,
                "total": total,
                "p0_total": p0_total,
                "mapped": mapped,
                "mapped_p0": mapped_p0,
                "coverage_ratio": round(coverage_ratio, 4),
                "coverage_text": f"{mapped} / {total}",
                "comment": comment,
            }
        )

    remaining_all = [
        {
            "id": item["id"],
            "name": item["name"],
            "dimension": item["dimension"],
            "frequency": item["frequency"],
            "priority": item["priority"],
            "frameworks": item["frameworks"],
            "note": item["note"],
        }
        for item in sorted(catalog.values(), key=lambda current: (current["priority"], current["order"]))
        if item["id"] not in available_set
    ]

    missing_priority = [
        {
            "id": item["id"],
            "name": item["name"],
            "dimension": item["dimension"],
            "frequency": item["frequency"],
            "priority": item["priority"],
            "frameworks": item["frameworks"],
            "note": item["note"],
        }
        for item in sorted(catalog.values(), key=lambda current: (current["priority"], current["order"]))
        if item["id"] not in available_set
        and (item["priority"] == "P0" or item["id"] in MISSING_PRIORITY_IDS)
    ][:18]
    missing_priority_ids = {item["id"] for item in missing_priority}
    remaining_optional = [item for item in remaining_all if item["id"] not in missing_priority_ids]

    summary = (
        "目前覆盖最完整的是货币、外部、商品和市场日频资产价格；"
        "增长、通胀、信用和库存硬数据大多仍待月/季频导出。"
        if monthly_status["status"] != "ready"
        else "当前已具备日频与月/季频双层覆盖，可同时观察价格信号和总量确认。"
    )
    return {
        "summary": summary,
        "dimensions": dimensions,
        "missing_priority": missing_priority,
        "remaining_all": remaining_optional,
    }


def build_highlights(signal_map: Dict[str, dict], monthly_status: dict) -> List[str]:
    highlights = []
    pmi = signal_map.get("PMI_MFG")
    social_fin = signal_map.get("SOCIAL_FIN_STOCK_YoY")
    dr007 = signal_map.get("DR007")
    cgb = signal_map.get("CGB_10Y")
    dxy = signal_map.get("DXY")
    gold = signal_map.get("COMEX_GOLD")
    csi = signal_map.get("CSI300")
    bond = signal_map.get("CBA_TOTAL_WEALTH")
    if pmi:
        highlights.append(f"制造业 PMI 当前 {pmi['latest_value']}，{pmi['signal_label']}。")
    if social_fin:
        highlights.append(f"社融存量同比当前 {social_fin['latest_value']}，用于确认信用修复是否持续。")
    if dr007 and cgb:
        highlights.append(f"短端资金利率与长端国债收益率同时处低位，显示货币环境仍偏友好。")
    if dxy:
        highlights.append(f"美元指数当前 {dxy['latest_value']}，外部流动性压力较前期缓和。")
    if gold:
        highlights.append(f"黄金位于 {gold['percentile_text']}，避险与再通胀对冲需求都在抬升。")
    if csi and bond:
        highlights.append(f"A 股与债券财富指数阶段性共振，但估值和信用硬数据仍需后验确认。")
    if monthly_status["status"] != "ready":
        highlights.append("增长、通胀、信用月/季频数据尚未落盘，当前看板更偏“价格先行信号板”。")
    return highlights[:5]


def build_summary_cards(catalog: Dict[str, dict], available_catalog_ids: Iterable[str], latest_sheet_date: datetime, monthly_status: dict) -> List[dict]:
    available_set = set(available_catalog_ids)
    total = len(catalog)
    p0_total = sum(1 for item in catalog.values() if item["priority"] == "P0")
    p0_mapped = sum(1 for item in catalog.values() if item["priority"] == "P0" and item["id"] in available_set)
    raw_source_mapped = len([item for item in available_set if item not in DERIVED_INDICATOR_IDS])
    derived_count = len([item for item in available_set if item in DERIVED_INDICATOR_IDS])
    p0_note = (
        "P0 核心指标已全部接入，日频价格信号与月/季频总量确认已打通。"
        if monthly_status["status"] == "ready" and p0_mapped == p0_total
        else "当前主要覆盖货币、外部、商品与市场日频资产价格。"
    )
    return [
        {
            "label": "数据更新时间",
            "value": iso_date(latest_sheet_date),
            "note": "以日度工作簿中的最新观测日期为准。",
        },
        {
            "label": "已接入指标",
            "value": f"{len(available_set)} / {total}",
            "note": f"含 {raw_source_mapped} 条原始序列和 {derived_count} 条派生指标。",
        },
        {
            "label": "P0核心覆盖",
            "value": f"{p0_mapped} / {p0_total}",
            "note": p0_note,
        },
        {
            "label": "月季频状态",
            "value": "待补" if monthly_status["status"] != "ready" else "已就绪",
            "note": monthly_status["summary"],
        },
    ]


def classify_growth_cycle(series_map: Dict[str, List[Tuple[datetime, float]]]) -> Tuple[str, str, List[str]]:
    pmi = latest_value(series_map, "PMI_MFG")
    new_orders = latest_value(series_map, "PMI_NEW_ORDER")
    gdp = latest_value(series_map, "GDP_YoY")
    iva = latest_value(series_map, "IVA_YoY")
    retail = latest_value(series_map, "RETAIL_YoY")

    score = 0
    evidence = []
    if pmi is not None:
        score += 1 if pmi >= 50 else -1
        evidence.append(f"制造业 PMI 为 {pmi:.1f}，{'处于扩张区间' if pmi >= 50 else '仍在收缩区间'}。")
    if new_orders is not None:
        score += 1 if new_orders >= 50 else -1
        evidence.append(f"新订单为 {new_orders:.1f}，{'需求边际回暖' if new_orders >= 50 else '需求仍偏弱'}。")
    if gdp is not None:
        if gdp >= 5:
            score += 1
        elif gdp < 4:
            score -= 1
        evidence.append(f"GDP 同比为 {gdp:.1f}%。")
    if iva is not None:
        if iva >= 5:
            score += 1
        elif iva < 4:
            score -= 1
        evidence.append(f"工业增加值同比为 {iva:.1f}%。")
    if retail is not None:
        if retail >= 3:
            score += 1
        elif retail < 0:
            score -= 1
        evidence.append(f"社零同比为 {retail:.1f}%。")

    growth_momentum = compare_momentum(
        safe_mean([pmi, new_orders, iva]),
        safe_mean(
            [
                previous_value(series_map, "PMI_MFG", 3),
                previous_value(series_map, "PMI_NEW_ORDER", 3),
                previous_value(series_map, "IVA_YoY", 3),
            ]
        ),
        0.4,
    )

    if score >= 4:
        return "增长修复", growth_momentum, evidence
    if score >= 1:
        return "弱修复", growth_momentum, evidence
    if score <= -3:
        return "增长下行", growth_momentum, evidence
    return "低位震荡", growth_momentum, evidence


def classify_inflation_cycle(series_map: Dict[str, List[Tuple[datetime, float]]]) -> Tuple[str, str, List[str]]:
    cpi = latest_value(series_map, "CPI_YoY")
    ppi = latest_value(series_map, "PPI_YoY")
    core_cpi = latest_value(series_map, "CORE_CPI_YoY")
    commodity = latest_value(series_map, "NH_COMMODITY")

    evidence = []
    if cpi is not None:
        evidence.append(f"CPI 同比为 {cpi:.1f}%。")
    if core_cpi is not None:
        evidence.append(f"核心 CPI 同比为 {core_cpi:.1f}%。")
    if ppi is not None:
        evidence.append(f"PPI 同比为 {ppi:.1f}%。")
    if commodity is not None:
        evidence.append(f"南华商品指数最新为 {commodity:.1f}。")

    inflation_level = "中性"
    if cpi is not None and ppi is not None:
        if cpi <= 1.2 and ppi <= 1.0:
            inflation_level = "低通胀"
        elif cpi >= 2.5 or ppi >= 2.0:
            inflation_level = "再通胀抬头"

    inflation_momentum = compare_momentum(
        safe_mean([cpi, ppi]),
        safe_mean([previous_value(series_map, "CPI_YoY", 3), previous_value(series_map, "PPI_YoY", 3)]),
        0.2,
    )
    return inflation_level, inflation_momentum, evidence


def classify_liquidity_state(series_map: Dict[str, List[Tuple[datetime, float]]]) -> Tuple[str, List[str]]:
    dr007 = latest_value(series_map, "DR007")
    omo = latest_value(series_map, "OMO_7D")
    cgb_10y = latest_value(series_map, "CGB_10Y")
    rrr = latest_value(series_map, "RRR_BIG")
    evidence = []
    if dr007 is not None and omo is not None:
        evidence.append(f"DR007 为 {dr007:.2f}%，7 天逆回购利率为 {omo:.2f}%。")
    if cgb_10y is not None:
        evidence.append(f"中国 10Y 国债收益率为 {cgb_10y:.2f}%。")
    if rrr is not None:
        evidence.append(f"大型金融机构存准率为 {rrr:.1f}%。")

    if dr007 is not None and omo is not None and dr007 <= omo + 0.15 and (cgb_10y or 99) <= 2.3:
        return "偏宽松", evidence
    if dr007 is not None and omo is not None and dr007 >= omo + 0.4:
        return "偏紧", evidence
    return "中性", evidence


def classify_credit_state(series_map: Dict[str, List[Tuple[datetime, float]]]) -> Tuple[str, List[str]]:
    social_fin = latest_value(series_map, "SOCIAL_FIN_STOCK_YoY")
    loan_bal = latest_value(series_map, "LOAN_BAL_YoY")
    corp_new = latest_value(series_map, "CORP_MLT_LOAN_NEW")
    corp_new_prev = previous_value(series_map, "CORP_MLT_LOAN_NEW", 12)
    m1_m2_gap = latest_value(series_map, "M1_M2_GAP")
    aaa_spread = latest_value(series_map, "CREDIT_SPREAD_AAA_3Y")

    score = 0
    evidence = []
    if social_fin is not None:
        if social_fin >= 7.5:
            score += 1
        elif social_fin < 7:
            score -= 1
        evidence.append(f"社融存量同比为 {social_fin:.1f}%。")
    if loan_bal is not None:
        if loan_bal >= 8:
            score += 1
        elif loan_bal < 6:
            score -= 1
        evidence.append(f"贷款余额同比为 {loan_bal:.1f}%。")
    if corp_new is not None:
        if corp_new >= 10000:
            score += 1
        if corp_new_prev is not None and corp_new > corp_new_prev:
            score += 1
            evidence.append(f"企业中长贷新增较去年同期改善至 {corp_new:,.0f} 亿元。")
        else:
            evidence.append(f"企业中长贷新增为 {corp_new:,.0f} 亿元。")
    if m1_m2_gap is not None:
        if m1_m2_gap > 0:
            score += 1
        elif m1_m2_gap < -4:
            score -= 1
        evidence.append(f"M1-M2 剪刀差为 {m1_m2_gap:.1f}。")
    if aaa_spread is not None:
        if aaa_spread <= 25:
            score += 1
        elif aaa_spread >= 60:
            score -= 1
        evidence.append(f"AAA 信用利差为 {aaa_spread:.1f}bp。")

    if score >= 3:
        return "宽信用扩张", evidence
    if score >= 1:
        return "信用修复早期", evidence
    return "信用偏弱", evidence


def classify_inventory_cycle(series_map: Dict[str, List[Tuple[datetime, float]]]) -> Tuple[str, str, List[str]]:
    pmi_new = latest_value(series_map, "PMI_NEW_ORDER")
    pmi_inv = latest_value(series_map, "PMI_FG_INV")
    inv = latest_value(series_map, "FG_INV_YTD_YoY")
    revenue = recent_indicator(series_map, "IND_REVENUE_YTD") or recent_indicator(series_map, "IND_MAIN_REV_YTD")

    def map_phase(demand_value: Optional[float], inventory_value: Optional[float], threshold: float = 0) -> str:
        if demand_value is None or inventory_value is None:
            return "信号不足"
        if demand_value > threshold and inventory_value <= threshold:
            return "被动去库"
        if demand_value > threshold and inventory_value > threshold:
            return "主动补库"
        if demand_value <= threshold and inventory_value > threshold:
            return "被动补库"
        return "主动去库"

    high_frequency_phase = map_phase(
        None if pmi_new is None else pmi_new - 50,
        None if pmi_inv is None else pmi_inv - 50,
        0,
    )
    industrial_phase = map_phase(revenue, inv, 0)
    evidence = []
    if pmi_new is not None and pmi_inv is not None:
        evidence.append(f"PMI 新订单 {pmi_new:.1f}，产成品库存 {pmi_inv:.1f}。")
    if inv is not None and revenue is not None:
        evidence.append(f"工业企业产成品库存同比 {inv:.1f}%，收入同比 {revenue:.1f}%。")

    if high_frequency_phase == industrial_phase:
        summary = high_frequency_phase
    else:
        summary = f"高频{high_frequency_phase}，工业口径{industrial_phase}"
    return summary, high_frequency_phase, evidence


def infer_growth_inflation_scenario(
    growth_momentum: str,
    inflation_momentum: str,
    inflation_level: str,
    cpi: Optional[float],
    ppi: Optional[float],
) -> str:
    if inflation_level == "低通胀" and (cpi or 99) <= 1.2 and (ppi or 99) <= 1.0:
        return "growth-up-inflation-down" if growth_momentum == "up" else "growth-down-inflation-down"
    if growth_momentum == "up" and inflation_momentum == "up":
        return "growth-up-inflation-up"
    if growth_momentum == "up" and inflation_momentum != "up":
        return "growth-up-inflation-down"
    if growth_momentum != "up" and inflation_momentum == "up":
        return "growth-down-inflation-up"
    return "growth-down-inflation-down"


def infer_money_credit_scenario(liquidity_state: str, credit_state: str) -> str:
    is_wide_money = liquidity_state == "偏宽松"
    is_wide_credit = credit_state in {"宽信用扩张", "信用修复早期"}
    if is_wide_money and is_wide_credit:
        return "wide-money-wide-credit"
    if is_wide_money and not is_wide_credit:
        return "wide-money-tight-credit"
    if not is_wide_money and is_wide_credit:
        return "tight-money-wide-credit"
    return "tight-money-tight-credit"


def build_data_driven_view(
    series_map: Dict[str, List[Tuple[datetime, float]]],
    latest_sheet_date: datetime,
) -> dict:
    growth_state, growth_momentum, growth_evidence = classify_growth_cycle(series_map)
    inflation_state, inflation_momentum, inflation_evidence = classify_inflation_cycle(series_map)
    liquidity_state, liquidity_evidence = classify_liquidity_state(series_map)
    credit_state, credit_evidence = classify_credit_state(series_map)
    inventory_state, inventory_high_frequency, inventory_evidence = classify_inventory_cycle(series_map)

    cpi = latest_value(series_map, "CPI_YoY")
    ppi = latest_value(series_map, "PPI_YoY")
    growth_inflation_scenario_id = infer_growth_inflation_scenario(
        growth_momentum,
        inflation_momentum,
        inflation_state,
        cpi,
        ppi,
    )
    money_credit_scenario_id = infer_money_credit_scenario(liquidity_state, credit_state)

    headline = (
        f"数据驱动判断更接近“{growth_state}、{inflation_state}、{liquidity_state}、{credit_state}”的组合，"
        f"对应情景上更像“{SCENARIO_TITLES[growth_inflation_scenario_id]} + {SCENARIO_TITLES[money_credit_scenario_id]}”。"
    )
    summary = (
        "当前高频与总量数据共同指向“增长边际修复，但信用与实体确认仍不完全同步”的阶段。"
        if credit_state in {"信用修复早期", "信用偏弱"}
        else "当前数据更偏向总量与信用共振修复，可逐步提高对风险资产和信用资产的跟踪优先级。"
    )

    scenario_note = (
        "景气和新订单已回到扩张区间，而通胀仍处低位，因此美林时钟更偏复苏象限。"
        if growth_inflation_scenario_id == "growth-up-inflation-down"
        else "增长与通胀动能尚未完全同步，当前仍需结合信用与库存信号做交叉验证。"
    )
    money_credit_note = (
        "资金利率贴近政策利率、信用利差明显收敛，但 M1 活化和贷款增速仍说明信用扩张尚在早期。"
        if money_credit_scenario_id == "wide-money-wide-credit"
        else "流动性与信用尚未形成完全共振，组合上仍需保留防守底仓。"
    )
    inventory_note = (
        "PMI 口径仍更像被动去库尾段，而工业企业库存和收入同比已经接近补库状态，说明库存周期正从高频验证走向实体确认。"
        if "高频" in inventory_state
        else "库存口径较一致，可直接作为库存周期位置的辅助判断。"
    )

    evidence = (
        growth_evidence[:3]
        + inflation_evidence[:2]
        + liquidity_evidence[:2]
        + credit_evidence[:2]
        + inventory_evidence[:1]
    )
    implications = [
        {
            "asset": "中国利率债",
            "view": "底仓保留",
            "reason": "低通胀和偏宽流动性仍支持久期资产，且实体确认尚未完全落地。",
        },
        {
            "asset": "中国股票",
            "view": "结构进攻",
            "reason": "增长边际改善更适合提升结构性风险预算，而不是一步切到全面进攻。",
        },
        {
            "asset": "中国信用债",
            "view": "逐步上调",
            "reason": "信用利差收敛和企业中长贷改善支持信用修复，但仍需观察 M1 活化。",
        },
        {
            "asset": "中国商品/黄金",
            "view": "黄金优先",
            "reason": "商品再通胀信号存在，但工业品仍需等待更强的信用与补库确认。",
        },
    ]

    return {
        "as_of_date": iso_date(latest_sheet_date),
        "headline": headline,
        "summary": summary,
        "regime_tags": [growth_state, inflation_state, liquidity_state, credit_state, inventory_state],
        "cycle_cards": [
            {
                "label": "增长周期",
                "value": growth_state,
                "note": "景气和总量数据共同显示经济处于修复但并非强复苏阶段。",
            },
            {
                "label": "美林时钟",
                "value": SCENARIO_TITLES[growth_inflation_scenario_id],
                "note": scenario_note,
                "scenario_id": growth_inflation_scenario_id,
                "scenario_title": SCENARIO_TITLES[growth_inflation_scenario_id],
            },
            {
                "label": "货币信用",
                "value": SCENARIO_TITLES[money_credit_scenario_id] + ("早期" if money_credit_scenario_id == "wide-money-wide-credit" and credit_state == "信用修复早期" else ""),
                "note": money_credit_note,
                "scenario_id": money_credit_scenario_id,
                "scenario_title": SCENARIO_TITLES[money_credit_scenario_id],
            },
            {
                "label": "库存周期",
                "value": inventory_state,
                "note": inventory_note,
            },
        ],
        "evidence": evidence,
        "implications": implications,
    }


def build_source_status(kind: str, path: Optional[Path], inspection: dict, parsed_series: Dict[str, List[Tuple[datetime, float]]]) -> dict:
    latest_date = None
    if parsed_series:
        latest_date = max(series[-1][0] for series in parsed_series.values() if series)
    label_map = {
        "daily": "日频工作簿",
        "monthly": "月/季频工作簿",
        "manual": "手工覆盖文件",
    }
    return {
        "kind": kind,
        "label": label_map.get(kind, kind),
        "file_name": path.name if path else "",
        "status": inspection["status"],
        "layout": inspection.get("layout", ""),
        "summary": inspection["summary"],
        "details": inspection.get("details", []),
        "parsed_series_count": len(parsed_series),
        "latest_date": iso_date(latest_date) if latest_date else "",
    }


def build_dashboard_data() -> dict:
    catalog = read_catalog()
    alias_map = build_catalog_alias_map(catalog)
    daily_file = resolve_latest_workbook(DAILY_GLOB)
    monthly_file = resolve_latest_workbook(MONTHLY_GLOB)
    lme_copper_file = resolve_latest_workbook(LME_COPPER_GLOB)
    daily_status = inspect_source_workbook(daily_file)
    monthly_status = inspect_source_workbook(monthly_file)
    lme_copper_status = inspect_source_workbook(lme_copper_file)

    daily_series, daily_latest = read_workbook_series(
        daily_file,
        daily_status,
        alias_map=alias_map,
        explicit_name_map={**SERIES_NAME_MAP, **SUPPORT_SERIES},
    )
    monthly_series, monthly_latest = read_workbook_series(
        monthly_file,
        monthly_status,
        alias_map=alias_map,
    )
    lme_copper_series, lme_copper_latest = read_workbook_series(
        lme_copper_file,
        lme_copper_status,
        alias_map=alias_map,
    )
    manual_series, manual_status = read_manual_override_series(MANUAL_OVERRIDE_FILE, catalog)

    # Manual overrides act as fallback when workbook exports are missing.
    # Once an indicator is available in Excel, prefer the workbook source.
    series_map = {**manual_series}
    for indicator_id, series in daily_series.items():
        series_map[indicator_id] = series
    for indicator_id, series in monthly_series.items():
        series_map[indicator_id] = series
    for indicator_id, series in lme_copper_series.items():
        series_map[indicator_id] = series

    derived_series = build_derived_series(series_map)
    series_map.update(derived_series)
    manual_latest = None
    if manual_series:
        manual_latest = max(series[-1][0] for series in manual_series.values() if series)
    latest_candidates = [item for item in [daily_latest, monthly_latest, lme_copper_latest, manual_latest] if item]
    if not latest_candidates:
        raise SystemExit("未读取到任何有效宏观时间序列。")
    latest_sheet_date = max(latest_candidates)

    signal_map = {}
    for indicator_id in INDICATOR_META:
        payload = build_signal_payload(indicator_id, catalog, series_map, latest_sheet_date)
        if payload:
            signal_map[indicator_id] = payload

    theme_payloads = []
    available_catalog_ids = {indicator_id for indicator_id in series_map if indicator_id in catalog}
    for theme in THEME_DEFINITIONS:
        signals = [signal_map[indicator_id] for indicator_id in theme["indicators"] if indicator_id in signal_map]
        if not signals:
            continue
        theme_payloads.append(
            {
                "id": theme["id"],
                "title": theme["title"],
                "summary_hint": theme["summary_hint"],
                "summary": theme_summary(theme["id"], signal_map),
                "source_ids": theme["source_ids"],
                "signals": signals,
            }
        )

    coverage = build_coverage(catalog, available_catalog_ids, monthly_status)
    source_status = [
        build_source_status("daily", daily_file, daily_status, daily_series),
        build_source_status("monthly", monthly_file, monthly_status, monthly_series),
        build_source_status("lme-copper", lme_copper_file, lme_copper_status, lme_copper_series),
        build_source_status("manual", MANUAL_OVERRIDE_FILE if MANUAL_OVERRIDE_FILE.exists() else None, manual_status, manual_series),
    ]
    data_driven_view = build_data_driven_view(series_map, latest_sheet_date)
    quality_notes = [
        "说明表明确指出春节效应、社融口径调整、核心 CPI 起始时间、LPR 改革、MLF 与 DR007 起始时间等问题，回测与同比分析不能机械处理。",
        f"当前识别到的日频文件为 `{daily_file.name if daily_file else '未找到'}`。",
        (
            f"当前识别到的月/季频文件为 `{monthly_file.name}`，状态为：{monthly_status['summary']}"
            if monthly_file
            else "当前未识别到月/季频工作簿。"
        ),
        (
            f"当前识别到的 LME 铜文件为 `{lme_copper_file.name}`，状态为：{lme_copper_status['summary']}"
            if lme_copper_file
            else "当前未识别到 LME 铜单独工作簿。"
        ),
        monthly_status["summary"],
        *monthly_status["details"],
        (
            f"手工覆盖文件状态：{manual_status['summary']}"
            if manual_status["status"] != "missing"
            else "如需补齐未出现在 Excel 中的指标，可在 `外部宏观数据/manual_macro_series_overrides.json` 中手工追加。"
        ),
        "系统已支持宽表时间序列和长表格式；只要月/季频文件真正写入值，就会自动接入增长、通胀、信用和库存信号。",
    ]

    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "generator": {
            "script": "scripts/build_macro_dashboard_data.py",
            "catalog_file": CATALOG_FILE.name,
            "daily_file": daily_file.name if daily_file else "",
            "monthly_file": monthly_file.name if monthly_file else "",
        },
        "latest_sheet_date": iso_date(latest_sheet_date),
        "summary_cards": build_summary_cards(catalog, available_catalog_ids, latest_sheet_date, monthly_status),
        "highlights": build_highlights(signal_map, monthly_status),
        "data_driven_view": data_driven_view,
        "themes": theme_payloads,
        "coverage": coverage,
        "source_status": source_status,
        "quality_notes": quality_notes,
        "references": list(SOURCE_LIBRARY.values()),
    }


def main() -> None:
    payload = build_dashboard_data()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print("已生成:")
    print(OUTPUT_JSON)


if __name__ == "__main__":
    main()
